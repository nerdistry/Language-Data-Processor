"""
    This file generates a list of translations from en-xx based on the
    'train' partition of each XLSX file in the specified directory.
 """
import os
import json
from openpyxl import load_workbook
from zipfile import BadZipFile
import logging

logging.basicConfig(level=logging.INFO)

def generate_translations_from_en_xx(processed_files_dir):
    """ Generates all translations from en-xx for all train datasets. """
    files = os.listdir(processed_files_dir)
    results = []

    logging.info("Processing Files")

    for file in files:
        if not file.endswith('.xlsx'):
            continue

        file_path = os.path.join(processed_files_dir, file)

        try:
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active

            headers = [cell.value for cell in sheet[1]]

            id_col_idx = headers.index('id')
            utterance_col_idx = headers.index('utterance')
            partition_col_idx = headers.index('partition')

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[partition_col_idx] == "train":
                    results.append({"id": row[id_col_idx],
                                    "utterance": row[utterance_col_idx]})
        except BadZipFile:
            continue

    return results


processed_files_dir = 'processed-dataset'
translations = generate_translations_from_en_xx(processed_files_dir)

with open('all_translations.json', 'w', encoding='utf-8') as f:
    json.dump(translations, f, ensure_ascii=False, indent=4)

logging.info('Translations have been saved to all_translations.json')
